from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = ROOT / "cars.xlsx"
OUTPUT_DIR = ROOT / "supabase" / "imports"


def stringify(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def write_csv(path: Path, headers: list[str], rows: Iterable[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)


def build_customers_sheet(wb):
    ws = wb["customers"]
    headers = [
        "legacy_row_id",
        "created_at_raw",
        "branch_name",
        "employee_name",
        "customer_name",
        "phone",
        "requested_car",
        "payment_plan",
        "attachments_raw",
        "notes_raw",
        "status",
        "nickname",
        "address",
        "whatsapp_prefix",
        "next_follow_up_raw",
        "visit_count_raw",
        "last_contact_raw",
        "trade_in_raw",
    ]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        cells = list(row) + [None] * max(0, 17 - len(row))
        rows.append(
            [
                str(row_idx),
                stringify(cells[0]),
                stringify(cells[1]),
                stringify(cells[2]),
                stringify(cells[3]),
                stringify(cells[4]),
                stringify(cells[5]),
                stringify(cells[6]),
                stringify(cells[7]),
                stringify(cells[8]),
                stringify(cells[9]),
                stringify(cells[10]),
                stringify(cells[11]),
                stringify(cells[12]),
                stringify(cells[13]),
                stringify(cells[14]),
                stringify(cells[15]),
                stringify(cells[16]),
            ]
        )
    return headers, rows


def build_inventory_sheet(wb):
    ws = wb["inventory"]
    headers = [
        "legacy_row_id",
        "model",
        "owner_name",
        "deal_type",
        "chassis_no",
        "condition_label",
        "availability_status",
        "price_raw",
        "gearbox",
        "fuel_type",
        "review_date_raw",
        "installment_price_raw",
        "color",
        "production_year_raw",
        "mileage_raw",
        "specs",
        "inspection",
        "photos_raw",
    ]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        cells = list(row) + [None] * max(0, 17 - len(row))
        rows.append(
            [
                str(row_idx),
                stringify(cells[0]),
                stringify(cells[1]),
                stringify(cells[2]),
                stringify(cells[3]),
                stringify(cells[4]),
                stringify(cells[5]),
                stringify(cells[6]),
                stringify(cells[7]),
                stringify(cells[8]),
                stringify(cells[9]),
                stringify(cells[10]),
                stringify(cells[11]),
                stringify(cells[12]),
                stringify(cells[13]),
                stringify(cells[14]),
                stringify(cells[15]),
                stringify(cells[16]),
            ]
        )
    return headers, rows


def build_staff_sheet(wb):
    ws = wb["staff"]
    headers = [
        "legacy_row_id",
        "telegram_chat_id",
        "full_name",
        "branch_name",
        "role_raw",
        "password_raw",
    ]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        cells = list(row) + [None] * max(0, 5 - len(row))
        rows.append(
            [
                str(row_idx),
                stringify(cells[0]),
                stringify(cells[1]),
                stringify(cells[2]),
                stringify(cells[3]),
                stringify(cells[4]),
            ]
        )
    return headers, rows


def build_simple_sheet(wb, sheet_name: str, headers: list[str]):
    ws = wb[sheet_name]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        rows.append([str(row_idx)] + [stringify(cell) for cell in row[: len(headers)]])
    return ["legacy_row_id"] + headers, rows


def build_showrooms_sheet(wb):
    ws = wb["settings"]
    headers = ["legacy_row_id", "branch_name"]
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell not in (None, "") for cell in row):
            continue
        rows.append([str(row_idx), stringify(row[0])])
    return headers, rows


def main():
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)

    outputs = {
        "legacy_customers_import.csv": build_customers_sheet(wb),
        "legacy_inventory_import.csv": build_inventory_sheet(wb),
        "legacy_staff_import.csv": build_staff_sheet(wb),
        "legacy_showrooms_import.csv": build_showrooms_sheet(wb),
        "legacy_notifications_import.csv": build_simple_sheet(
            wb,
            "notifications",
            ["created_at_raw", "target_role_or_room", "message", "status"],
        ),
        "legacy_customer_logs_import.csv": build_simple_sheet(
            wb,
            "customer_logs",
            [
                "created_at_raw",
                "customer_row_id_raw",
                "customer_name",
                "actor_name",
                "action",
                "next_follow_up_raw",
                "details",
            ],
        ),
        "legacy_audit_logs_import.csv": build_simple_sheet(
            wb,
            "audit_log",
            [
                "created_at_raw",
                "actor_name",
                "actor_role",
                "action",
                "entity_type",
                "entity_row_id_raw",
                "details",
            ],
        ),
    }

    for file_name, (headers, rows) in outputs.items():
        write_csv(OUTPUT_DIR / file_name, headers, rows)
        print(f"Wrote {file_name} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
